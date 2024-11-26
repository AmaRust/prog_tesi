from llama_index.embeddings.huggingface import HuggingFaceEmbedding
from llama_index.core import Settings, SimpleDirectoryReader, VectorStoreIndex
from llama_index.readers.file import PyMuPDFReader
from llama_index.core.retrievers import VectorIndexRetriever
from llama_index.core.query_engine import RetrieverQueryEngine
from llama_index.core.postprocessor import SimilarityPostprocessor
from llama_index.core.node_parser import SentenceSplitter
from llama_index.core.schema import TextNode
from llama_index.vector_stores.faiss import FaissVectorStore
import faiss
import torch
from llama_index.core.vector_stores import VectorStoreQuery
from llama_index.core.node_parser import SemanticSplitterNodeParser
from transformers import AutoModelForCausalLM, AutoTokenizer, AutoModelForSequenceClassification, pipeline
from llama_index.core import Document
from datetime import datetime
from huggingface_hub import login
from datasets import load_dataset

# hide the loading messages
import logging
import transformers
transformers.tokenization_utils.logger.setLevel(logging.ERROR)
transformers.configuration_utils.logger.setLevel(logging.ERROR)
transformers.modeling_utils.logger.setLevel(logging.ERROR)

# METRICS
import nltk
nltk.download('wordnet')
from nltk.translate.bleu_score import sentence_bleu
from rouge_score import rouge_scorer
from nltk.translate.meteor_score import meteor_score
from bert_score import score

import numpy as np
import random
import json
import os
import xml.etree.ElementTree as ET
from bs4 import BeautifulSoup

from openpyxl import Workbook

def get_choice(variable_name, options):
    print(f"Please choose a value for {variable_name} from the following options :")
    if variable_name not in ['splitting method', 'assessment method', 'reranking', 'documents reader', 'random pairs']:
        print("0. Enter your own value")
    for i, option in enumerate(options, 1):
        print(f"{i}. {option}")

    while True:
        try:
            choice = int(input(f"Enter the number of your choice for {variable_name} : "))
            if choice == 0:
                user_input = input(f"Enter your own value for {variable_name} : ")
                if options and all(isinstance(option, int) for option in options):
                    try:
                        user_input = int(user_input)
                    except ValueError:
                        print("Invalid entry. Please enter a valid integer.")
                        continue
                return user_input
            elif 1 <= choice <= len(options):
                return options[choice - 1]
            else:
                print(f"Invalid choice. Please choose a number between 0 and {len(options)}.")
        except ValueError:
            print("Invalid entry. Please enter a number.")

def count_tokens_transformers(text, model_name):
    tokenizer = AutoTokenizer.from_pretrained(model_name)
    tokens = tokenizer.tokenize(text)
    return len(tokens)

class XMLParentMap:
    def __init__(self, tree):
        self.tree = tree
        self.parent_map = self.build_parent_map()

    def build_parent_map(self):
        parent_map = {}
        for p in self.tree.iter():
            for c in p:
                if c in parent_map:
                    parent_map[c].append(p)
                else:
                    parent_map[c] = [p]
        return parent_map

    def get_all_parents(self, element):
        all_parents = []
        self._collect_all_parents(element, all_parents)
        return all_parents

    def _collect_all_parents(self, element, all_parents):
        if element in self.parent_map:
            parents = self.parent_map[element]
            for parent in parents:
                all_parents.append(parent)
                self._collect_all_parents(parent, all_parents)

class XMLDirectoryReader:
    def __init__(self, directory):
        self.directory = directory

    def load_data(self):
        documents = []
        for filename in os.listdir(self.directory):
            if filename.endswith(".xml"):
                file_path = os.path.join(self.directory, filename)
                with open(file_path, 'r', encoding='utf-8') as file:
                    tree = ET.parse(file)
                    root = tree.getroot()
                    parent_map_instance = XMLParentMap(tree)
                    text_content = self.extract_text_from_xml(root, parent_map_instance)
                    if text_content:  # Checks that the text is not empty
                        documents.append(text_content)
        return documents

    def extract_text_from_xml(self, root, parent_map):
        namespaces = {'tei': 'http://www.tei-c.org/ns/1.0'}
        texts = []
        for elem in root.iter():
            all_parents = parent_map.get_all_parents(elem)

            # Check whether the element or one of its parents is in <biblStruct>
            if any(p.tag in ['{http://www.tei-c.org/ns/1.0}biblStruct', '{http://www.tei-c.org/ns/1.0}figure'] for p in all_parents):
                continue  # Skip to next element if in <biblStruct>


            # Exclude certain specific elements (such as metadata or publication information)
            if elem.tag in ['{http://www.tei-c.org/ns/1.0}title', '{http://www.tei-c.org/ns/1.0}head', '{http://www.tei-c.org/ns/1.0}p']:
                if elem.text and elem.text.strip():  # Exclude elements with empty text
                    text = elem.text.strip()
                    if elem.tag == '{http://www.tei-c.org/ns/1.0}title':
                        text = f"\n\ntitle : {text}"  # Add the "title :" prefix
                    elif elem.tag == '{http://www.tei-c.org/ns/1.0}head':
                        text = f"\nhead : {text}"  # Add the "head :" prefix
                    texts.append(text)
        return "\n".join(texts)
    
    
class HTMLDirectoryReader:
    def __init__(self, directory):
        self.directory = directory

    def load_data(self):
        documents = []
        for filename in os.listdir(self.directory):
            if filename.endswith(".html") or filename.endswith(".htm"):
                file_path = os.path.join(self.directory, filename)
                with open(file_path, 'r', encoding='utf-8') as file:
                    soup = BeautifulSoup(file, 'lxml')
                    self.remove_references(soup) 
                    text_content = self.extract_text_from_html(soup)
                    if text_content:  
                        documents.append(text_content)
        return documents

    def remove_references(self, soup):
        references_section = soup.find('h2', id='references')
        if references_section:
            next_sibling = references_section.find_next_sibling()
            references_section.decompose()
            while next_sibling :
                next_sibling_to_decompose = next_sibling
                next_sibling = next_sibling.find_next_sibling()
                next_sibling_to_decompose.decompose()

    def extract_text_from_html(self, soup):
        text = soup.get_text()
        
        return text

# query2doc method for query expansion    
def query2doc(query, k, llm, tokenizer):
    dataset = load_dataset("intfloat/query2doc_msmarco")
    train_dataset = dataset['train']
    sampled_train_dataset = train_dataset.shuffle(seed=42).select(range(k))
    
    passages = "\n".join([f"Query: {sampled_train_dataset[i]['query']}\nPassage: {sampled_train_dataset[i]['pseudo_doc']}\n" for i in range(k)])
    
    prompt = f'''[INST] 
Write a passage that answers the given query (at the end):\n
{passages}
Query: {query}
Passage:  
[/INST]'''
    
    inputs = tokenizer(prompt, return_tensors="pt")

    outputs = llm.generate(input_ids=inputs["input_ids"].to("cuda"), max_new_tokens=500)

    generated_text = tokenizer.batch_decode(outputs, skip_special_tokens=True)[0]
    response = generated_text[len(prompt):].strip()
    
    new_query = f"{query} [SEP] {response}"

    return new_query


# chain of thought method 
def my_chain_of_thought(query, k, llm, tokenizer):
    dataset = load_dataset("medalpaca/medical_meadow_medical_flashcards")
    train_dataset = dataset['train']
    sampled_train_dataset = train_dataset.shuffle(seed=42).select(range(k))

    few_shot = "\n".join([f"Query: {sampled_train_dataset[i]['input']}\nAnswer: {sampled_train_dataset[i]['output']}\n" for i in range(k)])
    
    prompt = f'''[INST] 
For each answer, let's think step by step to complete them (in the answer area, replace the current text by your thinking):\n
{few_shot} 
[/INST]'''

    inputs = tokenizer(prompt, return_tensors="pt")

    outputs = llm.generate(input_ids=inputs["input_ids"].to("cuda"), max_new_tokens=1000)

    generated_text = tokenizer.batch_decode(outputs, skip_special_tokens=True)[0]
    response = generated_text[len(prompt):].strip()

    new_prompt = f'''[INST] 
{response}\n
---\n
Query: {query}
Answer: Let's think step by step to answer.
[/INST]'''
    
    inputs = tokenizer(new_prompt, return_tensors="pt")

    outputs = llm.generate(input_ids=inputs["input_ids"].to("cuda"), max_new_tokens=1000)

    generated_text = tokenizer.batch_decode(outputs, skip_special_tokens=True)[0]
    response = generated_text[len(new_prompt):].strip()

    return response

# reranking method to enhance the relevance of retrieved docs
def my_reranking(model_name, query, passages, top_n): # model_name : "cross-encoder/ms-marco-MiniLM-L-6-v2"
    tokenizer = AutoTokenizer.from_pretrained(model_name)
    model = AutoModelForSequenceClassification.from_pretrained(model_name)

    features = tokenizer([query] * len(passages), passages,  padding=True, truncation=True, return_tensors="pt")

    model.eval()
    with torch.no_grad():
        scores = model(**features).logits

    scores = scores.squeeze().tolist()

    scored_passages = list(zip(passages, scores))

    sorted_passages = sorted(scored_passages, key=lambda x: x[1], reverse=True)

    best_passages = []
    scores = []
    for passage, score in sorted_passages[:top_n]:
        best_passages.append(passage)
        scores.append(score)

    return (best_passages, scores)

    
def my_score_BLEU(reference, candidate):
    res = f"Individual 1-gram : {sentence_bleu(reference.split(), candidate, weights=(1, 0, 0, 0))}\nIndividual 2-gram : {sentence_bleu(reference.split(), candidate, weights=(0, 1, 0, 0))}\nIndividual 3-gram : {sentence_bleu(reference.split(), candidate, weights=(0, 0, 1, 0))}\nIndividual 4-gram : {sentence_bleu(reference.split(), candidate, weights=(0, 0, 0, 1))}"
    return res

def my_score_ROUGE(reference, candidate, scorer):
    scores = scorer.score(reference, candidate)
    res = f"ROUGE_1 : {scores['rouge1']}\nROUGE_L : {scores['rougeL']}"
    return res

def my_score_METEOR(reference, candidate):
    return meteor_score([reference.split()], candidate.split())

def my_BERTScore(reference, candidate):
    P, R, F1 = score([candidate], [reference], lang='en', rescale_with_baseline=True)
    res = f"Precision : {P.mean():.3f}\nRecall : {R.mean():.3f}\nF1 : {F1.mean():.3f}"
    return res 

def print_BLEU(reference, candidate, file):
    file.write(f"Individual 1-gram : {sentence_bleu(reference.split(), candidate, weights=(1, 0, 0, 0))}\n")
    file.write(f"Individual 2-gram : {sentence_bleu(reference.split(), candidate, weights=(0, 1, 0, 0))}\n")
    file.write(f"Individual 3-gram : {sentence_bleu(reference.split(), candidate, weights=(0, 0, 1, 0))}\n")
    file.write(f"Individual 4-gram : {sentence_bleu(reference.split(), candidate, weights=(0, 0, 0, 1))}\n\n")

def print_ROUGE(reference, candidate, file, scorer):
    scores = scorer.score(reference, candidate)
    file.write(f"ROUGE_1 : {scores['rouge1']}\n")
    file.write(f"ROUGE_L : {scores['rougeL']}\n\n")

def print_METEOR(reference, candidate, file):
    score = meteor_score([reference.split()], candidate.split())
    file.write(f"METEOR score : {score}\n\n")

def print_BERTScore(reference, candidate, file):
    # We obtain precision, recall and F1 scores for each sentence
    P, R, F1 = score([candidate], [reference], lang='en', verbose=True) 
    file.write(f"Precision : {P.mean():.3f}\n") 
    file.write(f"Recall : {R.mean():.3f}\n") 
    file.write(f"F1 : {F1.mean():.3f}\n\n") 

    # With "rescale_with_baseline=True" we can now see that the scores are much more spread out, which makes it easy to compare different examples.
    P, R, F1 = score([candidate], [reference], lang='en', rescale_with_baseline=True)
    file.write(f"Precision : {P.mean():.3f}\n") 
    file.write(f"Recall : {R.mean():.3f}\n") 
    file.write(f"F1 : {F1.mean():.3f}\n\n\n")

def print_mean_BLEU(references, candidates, file):
    ind_1_gram = [sentence_bleu(references[i].split(), candidates[i], weights=(1, 0, 0, 0)) for i in range(len(references))]
    ind_2_gram = [sentence_bleu(references[i].split(), candidates[i], weights=(0, 1, 0, 0)) for i in range(len(references))]
    ind_3_gram = [sentence_bleu(references[i].split(), candidates[i], weights=(0, 0, 1, 0)) for i in range(len(references))]
    ind_4_gram = [sentence_bleu(references[i].split(), candidates[i], weights=(0, 0, 0, 1)) for i in range(len(references))]
    file.write(f"Individual 1-gram : {np.mean(ind_1_gram)}\n")
    file.write(f"Individual 2-gram : {np.mean(ind_2_gram)}\n")
    file.write(f"Individual 3-gram : {np.mean(ind_3_gram)}\n")
    file.write(f"Individual 4-gram : {np.mean(ind_4_gram)}\n\n")

def print_mean_ROUGE(references, candidates, file, scorer):
    scores = [scorer.score(references[i], candidates[i]) for i in range(len(references))]

    precision_rouge_1 = [scores[i]['rouge1'][0] for i in range(len(scores))]
    recall_rouge_1 = [scores[i]['rouge1'][1] for i in range(len(scores))]
    fmeasure_rouge_1 = [scores[i]['rouge1'][2] for i in range(len(scores))]

    precision_rouge_L = [scores[i]['rougeL'][0] for i in range(len(scores))]
    recall_rouge_L = [scores[i]['rougeL'][1] for i in range(len(scores))]
    fmeasure_rouge_L = [scores[i]['rougeL'][2] for i in range(len(scores))]

    file.write(f"ROUGE_1 : precision={np.mean(precision_rouge_1)}, recall={np.mean(recall_rouge_1)}, fmeasure={np.mean(fmeasure_rouge_1)}\n")
    file.write(f"ROUGE_L : precision={np.mean(precision_rouge_L)}, recall={np.mean(recall_rouge_L)}, fmeasure={np.mean(fmeasure_rouge_L)}\n\n")

def print_mean_METEOR(references, candidates, file):
    scores = [meteor_score([references[i].split()], candidates[i].split()) for i in range(len(references))]
    file.write(f"METEOR score : {np.mean(scores)}\n\n")

def print_mean_BERTScore(references, candidates, file):
    scores = [score([candidates[i]], [references[i]], lang='en', verbose=True) for i in range(len(references))]
    
    P = [float(scores[i][0]) for i in range(len(scores))]
    R = [float(scores[i][1]) for i in range(len(scores))]
    F1 = [float(scores[i][2]) for i in range(len(scores))]
    
    file.write(f"Precision : {np.mean(P):.3f}\n") 
    file.write(f"Recall : {np.mean(R):.3f}\n") 
    file.write(f"F1 : {np.mean(F1):.3f}\n\n") 

    scores = [score([candidates[i]], [references[i]], lang='en', rescale_with_baseline=True) for i in range(len(references))]
    
    P = [float(scores[i][0]) for i in range(len(scores))]
    R = [float(scores[i][1]) for i in range(len(scores))]
    F1 = [float(scores[i][2]) for i in range(len(scores))]
    
    file.write(f"Precision : {np.mean(P):.3f}\n") 
    file.write(f"Recall : {np.mean(R):.3f}\n") 
    file.write(f"F1 : {np.mean(F1):.3f}\n\n") 


def load_pdfs_from_directory(directory_path):
    loader = PyMuPDFReader()
    documents = []
    
    for filename in os.listdir(directory_path):
        if filename.endswith(".pdf"):
            file_path = os.path.join(directory_path, filename)
            doc = loader.load(file_path=file_path)
            for page in doc :
                documents.append(page)
    
    return documents

def get_context_by_question(question, dataset):
        for entry in dataset:
            if entry["question"] == question:
                return entry["context"]
        return "Question not found."

def context_similarity(reference, candidate): 
    reference_modif = reference.replace('-', '')
    candidate_modif = candidate.replace('-', '')

    if reference_modif in candidate_modif:  # faire en sorte d'afficher le pourcentage du contexte de référence étant dans celui obtenu
        res = "----> The predicted context contains the reference context.\n\n"
    else:
        res = "----> The predicted context doesn't contain the reference context.\n\n"

    m_score = meteor_score([reference.split()], candidate.split())
    res += f"METEOR score : {m_score}\n\nBERTScore:\n"

    P, R, F1 = score([candidate], [reference], lang='en', rescale_with_baseline=True)
    res += f"Precision : {P.mean():.3f}\n"
    res += f"Recall : {R.mean():.3f}\n"
    res += f"F1 : {F1.mean():.3f}"

    return res

def contexts_similarity(references, candidates):
    scores = [meteor_score([references[i].split()], candidates[i].split()) for i in range(len(references))]
    res = f"METEOR score : {np.mean(scores)}\n\nBERTScore:\n"

    scores = [score([candidates[i]], [references[i]], lang='en', rescale_with_baseline=True) for i in range(len(references))]
    
    P = [float(scores[i][0]) for i in range(len(scores))]
    R = [float(scores[i][1]) for i in range(len(scores))]
    F1 = [float(scores[i][2]) for i in range(len(scores))]
    
    res += f"Precision : {np.mean(P):.3f}\n"
    res += f"Recall : {np.mean(R):.3f}\n" 
    res += f"F1 : {np.mean(F1):.3f}"

    return res

if __name__ == "__main__":

    login("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")

    # load the JSON file
    with open('../data/assessment_set/dataset.json', 'r') as dataset:
        qa_pairs = json.load(dataset)
    set_size = len(qa_pairs)

    # choice of parameters
    options_embedding_model = ["BAAI/bge-small-en-v1.5", "thenlper/gte-large", "sentence-transformers/all-MiniLM-L6-v2", "sentence-transformers/all-mpnet-base-v2", "nomic-ai/nomic-embed-text-v1.5", "Alibaba-NLP/gte-Qwen2-7B-instruct", "NeuML/pubmedbert-base-embeddings"]
    options_LLM_model = ["TheBloke/Mistral-7B-Instruct-v0.2-GPTQ", "microsoft/Phi-3-mini-128k-instruct", "TheBloke/Llama-2-7B-Chat-GGUF", "distilbert/distilgpt2", "tiiuae/falcon-7b-instruct", "meta-llama/Llama-2-7b-hf", "mistralai/Mistral-7B-Instruct-v0.2"]
    options_chunk_size = [128, 256, 512]
    options_chunk_overlap = [12, 25, 40]
    options_top_k = [1, 2, 3, 4, 5]
    options_query = ["What is the recommended targeted therapy for patients with BSI and severe infection due to 3GCephRE?",
                     "How is the evidence on treatment of infections caused by 3GCephRE organized?",
                     "Can imipenem-relebactam be recommended for the treatment of CRE infections?",
                     "Is there evidence for the use of eravacycline for the treatment of CRE infections?",
                     "Why does EUCAST not provide breakpoints for sulbactam and tigecycline for Acinetobacter?"]
    options_reference_response = ["A carbapenem (imipenem or meropenem) as targeted therapy",
                                  "The evidence on treatment of infections caused by 3GCephRE is organized by clinical syndrome, including bloodstream infections (BSI), urinary tract infections (UTI), pneumonia, and intra-abdominal infection (IAI).",
                                  "No, due to the paucity of data available for imipenem-relebactam against CRE, we cannot make recommendations on imipenem-relebactam for CRE at this time.",
                                  "No, there is no evidence for the use of eravacycline for the treatment of CRE infections, as no patients with CRE were included in the trials that led to its approval.",
                                  "EUCAST does not provide breakpoints for sulbactam and tigecycline for Acinetobacter, but the reason is not specified in the context."]
    options_max_new_tokens = [100, 190, 280, 370, 460, 550, 640, 730, 820, 910, 1000]
    options_splitting_method = ["Settings", "From scratch with Faiss vector store and SYNTACTIC splitter", "From scratch with Faiss vector store and SEMANTIC splitter"]
    options_test = ["Only one Q/A task", "Assessment of all the Q/A dataset", f"Assessment of a specific number of Q/A pairs (max : {set_size})"]
    options_reranking = ["Yes", "No"]
    options_top_n = [1, 2, 3]
    options_documents_reader = ["SimpleDirectoryReader for PDF", "PyMuPDFReader for PDF", "XML", "HTML"]
    options_random_pairs = ["Yes", "No"]

    documents_reader = get_choice('documents reader', options_documents_reader)
    embedding_model_name = get_choice('embedding model', options_embedding_model)
    LLM_model_name = get_choice('LLM model', options_LLM_model)
    splitting_method = get_choice('splitting method', options_splitting_method)
    if splitting_method != "From scratch with Faiss vector store and SEMANTIC splitter":
        chunk_size = get_choice('chunk size', options_chunk_size)
        chunk_overlap = get_choice('chunk overlap', options_chunk_overlap)
    top_k = get_choice('top_k value', options_top_k)
    test = get_choice('assessment method', options_test)
    if test == f"Assessment of a specific number of Q/A pairs (max : {len(qa_pairs)})":
        nb_qa_pairs = int(get_choice('number of Q/A pairs', []))
        random_pairs = get_choice('random pairs', options_random_pairs)
    elif test == "Only one Q/A task":
        query = get_choice('query', options_query)
        reference_response = get_choice('reference response', options_reference_response)
    max_new_tokens = get_choice('max new tokens', options_max_new_tokens)
    reranking = get_choice('reranking', options_reranking)
    if reranking == 'Yes':
        top_n = get_choice(f"top_n (max: {top_k})", options_top_n)
    

    # Print the parameters in the log file
    file = open(f'../experiment_logs/RAG/rag_{datetime.now().strftime("%Y-%m-%d %H:%M")}.txt', 'w')
    file.write(f"\n* Parameters for this experiment ({datetime.now().day}/{datetime.now().month}/{datetime.now().year}):\n\n")
    file.write(f"         - documents reader : {documents_reader}\n\n")
    file.write(f"         - embedding model : {embedding_model_name}\n\n")
    file.write(f"         - LLM model : {LLM_model_name}\n\n")
    file.write(f"         - splitting method : {splitting_method}\n\n")
    if splitting_method != "From scratch with Faiss vector store and SEMANTIC splitter":
        file.write(f"         - chunk size : {chunk_size}\n\n")
        file.write(f"         - chunk overlap : {chunk_overlap}\n\n")
    file.write(f"         - top_k value : {top_k}\n\n")
    file.write(f"         - assessment method : {test}\n\n")
    if test == "Only one Q/A task":
        file.write(f"         - query : {query}\n\n")
        file.write(f"         - reference response : {reference_response}\n\n")
    if test == f"Assessment of a specific number of Q/A pairs (max : {len(qa_pairs)})":
        file.write(f"         - number of Q/A pairs (random : {random_pairs}): {nb_qa_pairs}\n\n")   
    file.write(f"         - reranking : {reranking}\n\n")
    if reranking == "Yes":
        file.write(f"         - top_n : {top_n}\n\n")    
    file.write(f"         - max new tokens : {max_new_tokens}\n\n\n\n")
    file.write(f"*{'-'*50}*\n\n\n\n")

    #--------------------------------------------
    # Beginning of RAG

    # For testing with a specific number of Q/A pairs
    if test == f"Assessment of a specific number of Q/A pairs (max : {set_size})":
        if random_pairs == "No":
            qa_pairs = qa_pairs[:nb_qa_pairs] 
        else:
            qa_pairs = random.sample(qa_pairs, nb_qa_pairs)

    # load model
    model_name = LLM_model_name
    model = AutoModelForCausalLM.from_pretrained(model_name,
                                                device_map="auto",
                                                trust_remote_code=False,
                                                revision="main")
                                                # device_map="cuda",
                                                #  torch_dtype="auto", 
                                                #  trust_remote_code=True,
                                                # )  

    # load tokenizer
    tokenizer = AutoTokenizer.from_pretrained(model_name, use_fast=True)
    
    # #--------------------------------------------

    if documents_reader == "XML":
        documents = XMLDirectoryReader("../data/Medical_Guidelines/XML").load_data()

        # Transforming text strings into Document objects
        i = 1
        document_objects = [Document(text=doc, title=f"Document {i+1}") for i, doc in enumerate(documents)]

    elif documents_reader == "SimpleDirectoryReader for PDF":
        document_objects = SimpleDirectoryReader("../data/Medical_Guidelines/PDF").load_data()

    elif documents_reader == "PyMuPDFReader for PDF":
        document_objects = load_pdfs_from_directory("../data/Medical_Guidelines/PDF")

    elif documents_reader == "HTML":
        documents = HTMLDirectoryReader("../data/Medical_Guidelines/HTML").load_data()
        
        # Transforming text strings into Document objects
        i = 1
        document_objects = [Document(text=doc, title=f"Document {i+1}") for i, doc in enumerate(documents)]

    if test == "Only one Q/A task":
        reference_context = get_context_by_question(query, qa_pairs)

    else:
        reference_contexts = [pair['context'] for pair in qa_pairs]

    #--------------------------------------------

    if splitting_method == "Settings":
        Settings.embed_model = HuggingFaceEmbedding(model_name=embedding_model_name)

        Settings.llm = None
        Settings.chunk_size = chunk_size
        Settings.chunk_overlap = chunk_overlap

        

        # store docs into vector DB
        index = VectorStoreIndex.from_documents(document_objects)

        # configure retriever
        retriever = VectorIndexRetriever(
            index=index,
            similarity_top_k=top_k, # number of docs to retreive
        )

        # assemble query engine
        query_engine = RetrieverQueryEngine(
            retriever=retriever,
            node_postprocessors=[SimilarityPostprocessor(similarity_cutoff=0.5)],
        )

        if test == "Only one Q/A task":
            # query documents
            response = query_engine.query(query)

            # reformat response

            file.write(f"* {top_k} chunks retrieved for the query : \"{query}\"\n\n")

            context = "Context:\n"

            chunks = []
            for i in range(top_k):
                chunk_text = response.source_nodes[i].text
                chunks.append(chunk_text)     # useful if reranking
                context += f"{chunk_text}\n\n"
                file.write(f"- CHUNK {i + 1}:\n{chunk_text}\n\n")

            file.write(f"* GLOBAL CONTEXT:\n\n{context}\n\n")
            file.write(f"* REFERENCE CONTEXT:\n\n{reference_context}\n\n")
            file.write(f"* SIMILARITY:\n\n{context_similarity(reference_context, context)}\n\n")

            if reranking == "Yes":
                file.write("* AFTER RERANKING\n\n")
                context = "Context:\n"
                (new_chunks, scores) = my_reranking("cross-encoder/ms-marco-MiniLM-L-6-v2", query, chunks, top_n)

                for i in range(top_n):
                    context += f"{new_chunks[i]}\n\n"
                    file.write(f"- CHUNK {i + 1} (score: {scores[i]}) :\n{new_chunks[i]}\n\n")

                file.write(f"* GLOBAL CONTEXT:\n\n{context}\n\n")
                file.write(f"* REFERENCE CONTEXT:\n\n{reference_context}\n\n")
                file.write(f"* SIMILARITY:\n\n{context_similarity(reference_context, context)}\n\n")

            file.write(f"*{'-'*50}*\n\n\n\n")

        elif test == "Assessment of all the Q/A dataset" or test == f"Assessment of a specific number of Q/A pairs (max : {set_size})":
            contexts = []

            if reranking == "No":
                for qa_pair in qa_pairs:
                    response = query_engine.query(qa_pair['question'])
                    context = "Context:\n"
                    for i in range(top_k):
                        chunk_text = response.source_nodes[i].text
                        context += f"{chunk_text}\n\n"
                    contexts.append(context)

            else:
                for qa_pair in qa_pairs:
                    response = query_engine.query(qa_pair['question'])
                    context = "Context:\n"
                    chunks = []
                    for i in range(top_k):
                        chunk_text = response.source_nodes[i].text
                        chunks.append(chunk_text)
                    
                    (new_chunks, scores) = my_reranking("cross-encoder/ms-marco-MiniLM-L-6-v2", qa_pair['question'], chunks, top_n)
                    for i in range(top_n):
                        context += f"{new_chunks[i]}\n\n"
                    
                    contexts.append(context)

            file.write(f"* CONTEXTS SIMILARITY :\n\n{contexts_similarity(reference_contexts, contexts)}\n\n\n\n*{'-'*50}*\n\n\n\n")
           
    #--------------------------------------------

    elif splitting_method == "From scratch with Faiss vector store and SYNTACTIC splitter":

        embed_model = HuggingFaceEmbedding(model_name=embedding_model_name)

        text_parser = SentenceSplitter(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            # separator=" ",
        )

        text_chunks = []
        # maintain relationship with source doc index, to help inject doc metadata 
        doc_idxs = []
        for doc_idx, doc in enumerate(document_objects):
            cur_text_chunks = text_parser.split_text(doc.text)
            text_chunks.extend(cur_text_chunks)
            doc_idxs.extend([doc_idx] * len(cur_text_chunks))

        nodes = []
        for idx, text_chunk in enumerate(text_chunks):
            node = TextNode(
                text=text_chunk,
            )
            src_doc = document_objects[doc_idxs[idx]]
            node.metadata = src_doc.metadata
            nodes.append(node)

        for node in nodes:
            node_embedding = embed_model.get_text_embedding(
                node.get_content(metadata_mode="all")
            )
            node.embedding = node_embedding

        # create a faiss index
        d = len(nodes[0].embedding)  # dimension
        faiss_index = faiss.IndexFlatL2(d)

        vector_store = FaissVectorStore(faiss_index=faiss_index)

        vector_store.add(nodes)

        # vector_store.persist("./faiss_vector_store")

        # # Chemin où l'index FAISS sera sauvegardé
        # index_path = "./faiss_vector_store"

        # if os.path.exists(index_path):
        #     # Charger l'index existant
        #     faiss_index = faiss.read_index(index_path)
        #     print("Index FAISS chargé à partir du disque.")
        # else:
        #     # Créer un nouvel index FAISS
        #     d = len(nodes[0].embedding)  # dimension
        #     faiss_index = faiss.IndexFlatL2(d)

        #     vector_store = FaissVectorStore(faiss_index=faiss_index)

        #     # Ajouter les embeddings des nodes à l'index FAISS
        #     vector_store.add(nodes)

        #     # Sauvegarder l'index FAISS sur le disque
        #     faiss.write_index(faiss_index, index_path)
        #     print("Nouvel index FAISS créé et sauvegardé sur le disque.")

        # vector_store = FaissVectorStore(faiss_index=faiss_index)

        if test == "Only one Q/A task":
            query_str = query

            query_embedding = embed_model.get_query_embedding(query_str)

            query_mode = "default"
            # query_mode = "sparse"
            # query_mode = "hybrid"

            vector_store_query = VectorStoreQuery(
                query_embedding=query_embedding, similarity_top_k=top_k, mode=query_mode
            )

            # returns a VectorStoreQueryResult
            query_result = vector_store.query(vector_store_query)

            file.write(f"* {top_k} chunks retrieved for the query : \"{query}\"\n\n")

            context = "Context:\n"

            chunks = []
            for i in range(top_k):
                chunk_text = nodes[int(query_result.ids[top_k-i-1])].text
                chunks.append(chunk_text)   # useful if reranking
                context += f"{chunk_text}\n\n"
                file.write(f"- CHUNK {i + 1} (similarity : {query_result.similarities[top_k-i-1]}):\n{chunk_text}\n\n")


            file.write(f"* GLOBAL CONTEXT:\n\n{context}\n\n")
            file.write(f"* REFERENCE CONTEXT:\n\n{reference_context}\n\n")
            file.write(f"* SIMILARITY:\n\n{context_similarity(reference_context, context)}\n\n")

            if reranking == "Yes":
                file.write("* AFTER RERANKING\n\n")
                context = "Context:\n"
                (new_chunks, scores) = my_reranking("cross-encoder/ms-marco-MiniLM-L-6-v2", query, chunks, top_n)

                for i in range(top_n):
                    context += f"{new_chunks[i]}\n\n"
                    file.write(f"- CHUNK {i + 1} (score: {scores[i]}) :\n{new_chunks[i]}\n\n")

                file.write(f"* GLOBAL CONTEXT:\n\n{context}\n\n")
                file.write(f"* REFERENCE CONTEXT:\n\n{reference_context}\n\n")
                file.write(f"* SIMILARITY:\n\n{context_similarity(reference_context, context)}\n\n")


            file.write(f"*{'-'*50}*\n\n\n\n")

        elif test == "Assessment of all the Q/A dataset" or test == f"Assessment of a specific number of Q/A pairs (max : {set_size})":
            contexts = []

            query_mode = "default"
            # query_mode = "sparse"
            # query_mode = "hybrid"

            for qa_pair in qa_pairs:
                query_embedding = embed_model.get_query_embedding(qa_pair['question'])

                vector_store_query = VectorStoreQuery(
                    query_embedding=query_embedding, similarity_top_k=top_k, mode=query_mode
                )
                # returns a VectorStoreQueryResult
                query_result = vector_store.query(vector_store_query)

                context = "Context:\n"

                if reranking == "No":
                    for i in range(top_k):
                        chunk_text = nodes[int(query_result.ids[top_k-i-1])].text
                        context += f"{chunk_text}\n\n"
                    contexts.append(context)

                else:
                    chunks = []
                    for i in range(top_k):
                        chunk_text = nodes[int(query_result.ids[top_k-i-1])].text
                        chunks.append(chunk_text)
                    
                    (new_chunks, scores) = my_reranking("cross-encoder/ms-marco-MiniLM-L-6-v2", qa_pair['question'], chunks, top_n)
                    for i in range(top_n):
                        context += f"{new_chunks[i]}\n\n"
                    
                    contexts.append(context)

            file.write(f"* CONTEXTS SIMILARITY :\n\n{contexts_similarity(reference_contexts, contexts)}\n\n\n\n*{'-'*50}*\n\n\n\n")

    #--------------------------------------------

    elif splitting_method == "From scratch with Faiss vector store and SEMANTIC splitter":
        
        embed_model = HuggingFaceEmbedding(model_name=embedding_model_name)

        semantic_parser = SemanticSplitterNodeParser(
            buffer_size=1, breakpoint_percentile_threshold=85, embed_model=embed_model
        )

        nodes = semantic_parser.build_semantic_nodes_from_documents(document_objects)

        for node in nodes:
            node_embedding = embed_model.get_text_embedding(
                node.get_content(metadata_mode="all")
            )
            node.embedding = node_embedding

        # create a faiss index
        d = len(nodes[0].embedding)  # dimension
        faiss_index = faiss.IndexFlatL2(d)

        vector_store = FaissVectorStore(faiss_index=faiss_index)

        vector_store.add(nodes)

        #vector_store.persist("./faiss_vector_store")

        # index_path = "./faiss_vector_store"

        # if os.path.exists(index_path):
        #     # Charger l'index existant
        #     faiss_index = faiss.read_index(index_path)
        #     print("Index FAISS chargé à partir du disque.")
        # else:
        #     # Créer un nouvel index FAISS
        #     d = len(nodes[0].embedding)  # dimension
        #     faiss_index = faiss.IndexFlatL2(d)

        #     vector_store = FaissVectorStore(faiss_index=faiss_index)

        #     # Ajouter les embeddings des nodes à l'index FAISS
        #     vector_store.add(nodes)

        #     # Sauvegarder l'index FAISS sur le disque
        #     faiss.write_index(faiss_index, index_path)
        #     print("Nouvel index FAISS créé et sauvegardé sur le disque.")

        # vector_store = FaissVectorStore(faiss_index=faiss_index)

        if test == "Only one Q/A task":
            query_str = query

            query_embedding = embed_model.get_query_embedding(query_str)

            query_mode = "default"
            # query_mode = "sparse"
            # query_mode = "hybrid"

            vector_store_query = VectorStoreQuery(
                query_embedding=query_embedding, similarity_top_k=top_k, mode=query_mode
            )

            # returns a VectorStoreQueryResult
            query_result = vector_store.query(vector_store_query)

            file.write(f"* {top_k} chunks retrieved for the query : \"{query}\"\n\n")

            context = "Context:\n"

            chunks = []
            for i in range(top_k):
                chunk_text = nodes[int(query_result.ids[top_k-i-1])].text
                chunks.append(chunk_text)     # useful if reranking
                context += f"{chunk_text}\n\n"
                file.write(f"- CHUNK {i + 1} (similarity : {query_result.similarities[top_k-i-1]}):\n{chunk_text}\n\n")

            file.write(f"* GLOBAL CONTEXT:\n\n{context}\n\n")
            file.write(f"* REFERENCE CONTEXT:\n\n{reference_context}\n\n")
            file.write(f"* SIMILARITY:\n\n{context_similarity(reference_context, context)}\n\n")

            if reranking == 'Yes':
                file.write("* AFTER RERANKING\n\n")
                context = "Context:\n"
                (new_chunks, scores) = my_reranking("cross-encoder/ms-marco-MiniLM-L-6-v2", query, chunks, top_n)

                for i in range(top_n):
                    context += f"{new_chunks[i]}\n\n"
                    file.write(f"- CHUNK {i + 1} (score: {scores[i]}) :\n{new_chunks[i]}\n\n")

                file.write(f"* GLOBAL CONTEXT:\n\n{context}\n\n")
                file.write(f"* REFERENCE CONTEXT:\n\n{reference_context}\n\n")
                file.write(f"* SIMILARITY:\n\n{context_similarity(reference_context, context)}\n\n")

            file.write(f"*{'-'*50}*\n\n\n\n")

        elif test == "Assessment of all the Q/A dataset" or test == f"Assessment of a specific number of Q/A pairs (max : {set_size})":
            contexts = []

            query_mode = "default"
            # query_mode = "sparse"
            # query_mode = "hybrid"

            for qa_pair in qa_pairs:
                query_embedding = embed_model.get_query_embedding(qa_pair['question'])

                vector_store_query = VectorStoreQuery(
                    query_embedding=query_embedding, similarity_top_k=top_k, mode=query_mode
                )
                # returns a VectorStoreQueryResult
                query_result = vector_store.query(vector_store_query)

                context = "Context:\n"

                if reranking == 'No':
                    for i in range(top_k):
                        chunk_text = nodes[int(query_result.ids[top_k-i-1])].text
                        context += f"{chunk_text}\n\n"
                    contexts.append(context)

                else:
                    chunks = []
                    for i in range(top_k):
                        chunk_text = nodes[int(query_result.ids[top_k-i-1])].text
                        chunks.append(chunk_text)
                    
                    (new_chunks, scores) = my_reranking("cross-encoder/ms-marco-MiniLM-L-6-v2", qa_pair['question'], chunks, top_n)
                    for i in range(top_n):
                        context += f"{new_chunks[i]}\n\n"
                    
                    contexts.append(context)

            file.write(f"* CONTEXTS SIMILARITY :\n\n{contexts_similarity(reference_contexts, contexts)}\n\n\n\n*{'-'*50}*\n\n\n\n")

    #--------------------------------------------

    # prompt (no additional prompt and no context)

    if test == "Only one Q/A task":
        file.write(f"** NO ADDITIONAL PROMPT / NO CONTEXT **\n")

        prompt=f'''[INST] {query} [/INST]'''

        file.write(f"* The prompt we are going to use without additional prompt and context is :\n\n{prompt}\n\n")

        model.eval()

        inputs = tokenizer(prompt, return_tensors="pt")

        outputs = model.generate(input_ids=inputs["input_ids"].to("cuda"), max_new_tokens=max_new_tokens)

        generated_text = tokenizer.batch_decode(outputs, skip_special_tokens=True)[0]
        response1 = generated_text[len(prompt):].strip()

        file.write(f"* The answer obtained without additional prompt and context is :\n\n{response1}\n\n")
        file.write(f"* REFERENCE ANSWER :\n{reference_response}\n\n\n\n")
        file.write(f"*{'-'*50}*\n\n\n\n")

    elif test == "Assessment of all the Q/A dataset" or test == f"Assessment of a specific number of Q/A pairs (max : {set_size})":
        responses1 = []

        for qa_pair in qa_pairs:
            prompt=f'''[INST] {qa_pair['question']} [/INST]'''

            model.eval()

            inputs = tokenizer(prompt, return_tensors="pt")

            outputs = model.generate(input_ids=inputs["input_ids"].to("cuda"), max_new_tokens=max_new_tokens)

            generated_text = tokenizer.batch_decode(outputs, skip_special_tokens=True)[0]
            response1 = generated_text[len(prompt):].strip()
            responses1.append(response1)

    # prompt (additional prompt and no context)

    intstructions_string = f"""You are an expert medical professional with extensive knowledge in various fields of medicine, \n
        You are here to provide accurate, evidence-based information and guidance on a wide range of medical topics. \n
        Preferably give clear answers without too many details.

        Please respond to the following question.
        """
    prompt_template = lambda question: f'''[INST] {intstructions_string} \n{question} \n[/INST]'''

    if test == "Only one Q/A task":
        file.write(f"** ADDITIONAL PROMPT / NO CONTEXT **\n")

        prompt = prompt_template(query)
        file.write(f"* The prompt we are going to use with additional prompt but without context is :\n\n{prompt}\n\n")

        model.eval()

        inputs = tokenizer(prompt, return_tensors="pt")
        outputs = model.generate(input_ids=inputs["input_ids"].to("cuda"), max_new_tokens=max_new_tokens)

        generated_text = tokenizer.batch_decode(outputs, skip_special_tokens=True)[0]
        response2 = generated_text[len(prompt):].strip()

        file.write(f"* The answer obtained with additional prompt but without context is :\n\n{response2}\n\n")
        file.write(f"* REFERENCE ANSWER :\n{reference_response}\n\n\n\n")
        file.write(f"*{'-'*50}*\n\n\n\n")

    elif test == "Assessment of all the Q/A dataset" or test == f"Assessment of a specific number of Q/A pairs (max : {set_size})":
        responses2 = []

        for qa_pair in qa_pairs:
            prompt = prompt_template(qa_pair['question'])

            model.eval()

            inputs = tokenizer(prompt, return_tensors="pt")
            outputs = model.generate(input_ids=inputs["input_ids"].to("cuda"), max_new_tokens=max_new_tokens)

            generated_text = tokenizer.batch_decode(outputs, skip_special_tokens=True)[0]
            response2 = generated_text[len(prompt):].strip()
            responses2.append(response2)

    # prompt (no additional prompt and context)

    prompt_template_w_context = lambda context, question: f"""[INST]
    {context}
    Please respond to the following question. Use the context above if it is helpful.

    {question}
    [/INST]
    """

    if test == "Only one Q/A task":
        file.write(f"** NO ADDITIONAL PROMPT / CONTEXT **\n")

        prompt = prompt_template_w_context(context, query)
        file.write(f"* The prompt we are going to use with context but without additional prompt is :\n\n{prompt}\n\n")

        model.eval()

        inputs = tokenizer(prompt, return_tensors="pt")
        outputs = model.generate(input_ids=inputs["input_ids"].to("cuda"), max_new_tokens=max_new_tokens)

        generated_text = tokenizer.batch_decode(outputs, skip_special_tokens=True)[0]
        response3 = generated_text[len(prompt):].strip()

        file.write(f"* The answer obtained with context but without additional prompt is :\n\n{response3}\n\n")
        file.write(f"* REFERENCE ANSWER :\n{reference_response}\n\n\n\n")
        file.write(f"*{'-'*50}*\n\n\n\n")

    elif test == "Assessment of all the Q/A dataset" or test == f"Assessment of a specific number of Q/A pairs (max : {set_size})":
        responses3 = []

        for i, qa_pair in enumerate(qa_pairs):
            prompt = prompt_template_w_context(contexts[i], qa_pair['question'])

            model.eval()

            inputs = tokenizer(prompt, return_tensors="pt")
            outputs = model.generate(input_ids=inputs["input_ids"].to("cuda"), max_new_tokens=max_new_tokens)

            generated_text = tokenizer.batch_decode(outputs, skip_special_tokens=True)[0]
            response3 = generated_text[len(prompt):].strip()
            responses3.append(response3)

    # prompt (context and additional prompt)

    prompt_template_w_context = lambda context, question: f"""[INST]You are an expert medical professional with extensive knowledge in various fields of medicine, \n
    You are here to provide accurate, evidence-based information and guidance on a wide range of medical topics. \n
    Preferably give clear answers without too many details.

    {context}
    Please respond to the following question. Use the context above if it is helpful.

    {question}
    [/INST]
    """

    if test == "Only one Q/A task":
        file.write(f"** ADDITIONAL PROMPT / CONTEXT **\n")
        
        prompt = prompt_template_w_context(context, query)
        file.write(f"* The prompt we are going to use with context and additional prompt is :\n\n{prompt}\n\n")

        model.eval()

        inputs = tokenizer(prompt, return_tensors="pt")
        outputs = model.generate(input_ids=inputs["input_ids"].to("cuda"), max_new_tokens=max_new_tokens)

        generated_text = tokenizer.batch_decode(outputs, skip_special_tokens=True)[0]
        response4 = generated_text[len(prompt):].strip()

        file.write(f"* The answer obtained with context and additional prompt is :\n\n{response4}\n\n")
        file.write(f"* REFERENCE ANSWER :\n{reference_response}\n\n\n\n")

    elif test == "Assessment of all the Q/A dataset" or test == f"Assessment of a specific number of Q/A pairs (max : {set_size})":
        responses4 = []

        for i, qa_pair in enumerate(qa_pairs):
            prompt = prompt_template_w_context(contexts[i], qa_pair['question'])

            model.eval()

            inputs = tokenizer(prompt, return_tensors="pt")
            outputs = model.generate(input_ids=inputs["input_ids"].to("cuda"), max_new_tokens=max_new_tokens)

            generated_text = tokenizer.batch_decode(outputs, skip_special_tokens=True)[0]
            response4 = generated_text[len(prompt):].strip()
            responses4.append(response4)

    #--------------------------------------------

    # Use of metrics 
    scorer = rouge_scorer.RougeScorer(['rouge1', 'rougeL'], use_stemmer=True)

    if test == "Only one Q/A task":

        file.write(f"*{'-'*50}*\n\n\n\n")
        file.write("** RESULTS FOR EACH CASE WITH DIFFERENT METRICS **\n\n")

        file.write(f"* NO ADDITIONAL PROMPT / NO CONTEXT\n")
        file.write(f"         - BLEU : \n\n")
        print_BLEU(reference_response, response1, file)

        file.write(f"         - ROUGE : \n\n")
        print_ROUGE(reference_response, response1, file, scorer)

        file.write(f"         - METEOR : \n\n")
        print_METEOR(reference_response, response1, file)

        file.write(f"         - BERTScore : \n\n")
        print_BERTScore(reference_response, response1, file)

        file.write(f"* ADDITIONAL PROMPT / NO CONTEXT\n")
        file.write(f"         - BLEU : \n\n")
        print_BLEU(reference_response, response2, file)

        file.write(f"         - ROUGE : \n\n")
        print_ROUGE(reference_response, response2, file, scorer)

        file.write(f"         - METEOR : \n\n")
        print_METEOR(reference_response, response2, file)

        file.write(f"         - BERTScore : \n\n")
        print_BERTScore(reference_response, response2, file)

        file.write(f"* NO ADDITIONAL PROMPT / CONTEXT\n")
        file.write(f"         - BLEU : \n\n")
        print_BLEU(reference_response, response3, file)

        file.write(f"         - ROUGE : \n\n")
        print_ROUGE(reference_response, response3, file, scorer)

        file.write(f"         - METEOR : \n\n")
        print_METEOR(reference_response, response3, file)

        file.write(f"         - BERTScore : \n\n")
        print_BERTScore(reference_response, response3, file)

        file.write(f"* ADDITIONAL PROMPT / CONTEXT\n")
        file.write(f"         - BLEU : \n\n")
        print_BLEU(reference_response, response4, file)

        file.write(f"         - ROUGE : \n\n")
        print_ROUGE(reference_response, response4, file, scorer)

        file.write(f"         - METEOR : \n\n")
        print_METEOR(reference_response, response4, file)

        file.write(f"         - BERTScore : \n\n")
        print_BERTScore(reference_response, response4, file)

    elif test == "Assessment of all the Q/A dataset" or test == f"Assessment of a specific number of Q/A pairs (max : {set_size})":
        file.write("** RESULTS FOR EACH CASE WITH DIFFERENT METRICS **\n\n")

        reference_answers = [pair['answer'] for pair in qa_pairs]

        file.write(f"* NO ADDITIONAL PROMPT / NO CONTEXT\n")
        file.write(f"         - BLEU : \n\n")
        print_mean_BLEU(reference_answers, responses1, file)

        file.write(f"         - ROUGE : \n\n")
        print_mean_ROUGE(reference_answers, responses1, file, scorer)

        file.write(f"         - METEOR : \n\n")
        print_mean_METEOR(reference_answers, responses1, file)

        file.write(f"         - BERTScore : \n\n")
        print_mean_BERTScore(reference_answers, responses1, file)

        file.write(f"* ADDITIONAL PROMPT / NO CONTEXT\n")
        file.write(f"         - BLEU : \n\n")
        print_mean_BLEU(reference_answers, responses2, file)

        file.write(f"         - ROUGE : \n\n")
        print_mean_ROUGE(reference_answers, responses2, file, scorer)

        file.write(f"         - METEOR : \n\n")
        print_mean_METEOR(reference_answers, responses2, file)

        file.write(f"         - BERTScore : \n\n")
        print_mean_BERTScore(reference_answers, responses2, file)

        file.write(f"* NO ADDITIONAL PROMPT / CONTEXT\n")
        file.write(f"         - BLEU : \n\n")
        print_mean_BLEU(reference_answers, responses3, file)

        file.write(f"         - ROUGE : \n\n")
        print_mean_ROUGE(reference_answers, responses3, file, scorer)

        file.write(f"         - METEOR : \n\n")
        print_mean_METEOR(reference_answers, responses3, file)

        file.write(f"         - BERTScore : \n\n")
        print_mean_BERTScore(reference_answers, responses3, file)

        file.write(f"* ADDITIONAL PROMPT / CONTEXT\n")
        file.write(f"         - BLEU : \n\n")
        print_mean_BLEU(reference_answers, responses4, file)

        file.write(f"         - ROUGE : \n\n")
        print_mean_ROUGE(reference_answers, responses4, file, scorer)

        file.write(f"         - METEOR : \n\n")
        print_mean_METEOR(reference_answers, responses4, file)

        file.write(f"         - BERTScore : \n\n")
        print_mean_BERTScore(reference_answers, responses4, file)

        # Create a new workbook
        wb = Workbook()

        # Select the active sheet
        ws1 = wb.active
        ws1.title = "NO ADDITIONAL PROMPT -- NO CONTEXT"

        # Adding column headers
        headers = ["Question", "Reference answer", "Predicted answer", "BLEU score", "ROUGE score", "METEOR score", "BERTScore"]
        ws1.append(headers)

        data = []
        for i, qa_pair in enumerate(qa_pairs):
            data.append({"Question": qa_pair['question'], "Reference answer": qa_pair['answer'], "Predicted answer": responses1[i], "BLEU score": my_score_BLEU(qa_pair['answer'], responses1[i]), "ROUGE score": my_score_ROUGE(qa_pair['answer'], responses1[i], scorer), "METEOR score": my_score_METEOR(qa_pair['answer'], responses1[i]), "BERTScore": my_BERTScore(qa_pair['answer'], responses1[i])})
            
        for entry in data:
            ws1.append([entry["Question"], entry["Reference answer"], entry["Predicted answer"], entry["BLEU score"], entry["ROUGE score"], entry["METEOR score"], entry["BERTScore"]])

        ws2 = wb.create_sheet(title="ADDITIONAL PROMPT -- NO CONTEXT")

        ws2.append(headers)

        data = []
        for i, qa_pair in enumerate(qa_pairs):
            data.append({"Question": qa_pair['question'], "Reference answer": qa_pair['answer'], "Predicted answer": responses2[i], "BLEU score": my_score_BLEU(qa_pair['answer'], responses2[i]), "ROUGE score": my_score_ROUGE(qa_pair['answer'], responses2[i], scorer), "METEOR score": my_score_METEOR(qa_pair['answer'], responses2[i]), "BERTScore": my_BERTScore(qa_pair['answer'], responses2[i])})
            
        for entry in data:
            ws2.append([entry["Question"], entry["Reference answer"], entry["Predicted answer"], entry["BLEU score"], entry["ROUGE score"], entry["METEOR score"], entry["BERTScore"]])

        ws3 = wb.create_sheet(title="NO ADDITIONAL PROMPT -- CONTEXT")

        ws3.append(headers)

        data = []
        for i, qa_pair in enumerate(qa_pairs):
            data.append({"Question": qa_pair['question'], "Reference answer": qa_pair['answer'], "Predicted answer": responses3[i], "BLEU score": my_score_BLEU(qa_pair['answer'], responses3[i]), "ROUGE score": my_score_ROUGE(qa_pair['answer'], responses3[i], scorer), "METEOR score": my_score_METEOR(qa_pair['answer'], responses3[i]), "BERTScore": my_BERTScore(qa_pair['answer'], responses3[i])})
            
        for entry in data:
            ws3.append([entry["Question"], entry["Reference answer"], entry["Predicted answer"], entry["BLEU score"], entry["ROUGE score"], entry["METEOR score"], entry["BERTScore"]])

        ws4 = wb.create_sheet(title="ADDITIONAL PROMPT -- CONTEXT")

        ws4.append(headers)

        data = []
        for i, qa_pair in enumerate(qa_pairs):
            data.append({"Question": qa_pair['question'], "Reference answer": qa_pair['answer'], "Predicted answer": responses4[i], "BLEU score": my_score_BLEU(qa_pair['answer'], responses4[i]), "ROUGE score": my_score_ROUGE(qa_pair['answer'], responses4[i], scorer), "METEOR score": my_score_METEOR(qa_pair['answer'], responses4[i]), "BERTScore": my_BERTScore(qa_pair['answer'], responses4[i])})
            
        for entry in data:
            ws4.append([entry["Question"], entry["Reference answer"], entry["Predicted answer"], entry["BLEU score"], entry["ROUGE score"], entry["METEOR score"], entry["BERTScore"]])

        wb.save("../experiment_logs/RAG/results.xlsx")

        print("The Excel file has been successfully created.")


    # Close the output file
    file.close()